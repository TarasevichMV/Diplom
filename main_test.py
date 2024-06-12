import sqlite3
import pandas as pd
import numpy as np
from tensorflow.keras.models import load_model
from tensorflow.keras.layers import TextVectorization
from preprocessing.textPreprocessor import fasttext_lowercase_train_config, TextPreprocessor
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime
import re
import pickle
import tensorflow as tf

# Путь к базе данных и файлу Excel
db_path = 'mydatabase.db'
xlsx_path = 'news.xlsx'

# Сопоставление каналов и компаний
channel_to_company = {
    'tmk_group': 'ТМК',
    'evrazcom': 'Евраз',
    'uralsteel_official': 'ЗТЗ',
    'metalloinvest_official': 'Металлоинвест',
    'MMK_official': 'ММК',
    'nlmk_official': 'НЛМК',
    'severstal': 'Северсталь'
}

# Подключение к базе данных и извлечение сообщений
conn = sqlite3.connect(db_path)
query = "SELECT id, channel_username, message_id, message_text FROM messages WHERE analized = 0 LIMIT 50000"
df = pd.read_sql_query(query, conn)

# Инициализация препроцессора
PREPROCESSOR = TextPreprocessor(fasttext_lowercase_train_config)

# Предобработка сообщений
def preprocess(text):
    return PREPROCESSOR.text_cleaning(text)

df['processed_text'] = df['message_text'].apply(preprocess)

# Загрузка модели Keras
model_path = 'saved_model/my_model.keras'
model = load_model(model_path)
print(f"Модель загружена из '{model_path}'")

# Загрузка vectorizer
vectorizer_path = 'saved_model/vectorizer.pkl'
with open(vectorizer_path, 'rb') as f:
    vectorizer = pickle.load(f)
print(f"Vectorizer загружен из '{vectorizer_path}'")

# Прогон через модель и получение вероятностей
def get_probabilities(text):
    text_vectorized = vectorizer([text])
    probabilities = model.predict(text_vectorized)[0]
    return probabilities

# Получение заголовка из первого предложения
def extract_title(text):
    text = text.strip("*")
    match = re.match(r'^(.*?)([.!?])', text)
    if match:
        return match.group(1)
    return text.split()[0] if text else "No Title"

# Загрузка xlsx-файла
wb = load_workbook(xlsx_path)
sheet = wb['Свод новостей']  # Получаем нужный лист

# Определение начальной строки для записи
current_row = sheet.max_row + 1  # Находим первую пустую строку

style_name = 'datetime'
if style_name not in wb.named_styles:
    date_style = NamedStyle(name=style_name, number_format='MM/DD/YYYY')
    wb.add_named_style(date_style)

# Итерация по текстам и обработка
for index, row in df.iterrows():
    probabilities = get_probabilities(row['processed_text'])
    
    # Запись данных, если есть хоть одна вероятность больше 0.49
    high_probabilities = [prob for prob in probabilities if prob > 0.51]
    
    if high_probabilities:
        # Запись данных в xlsx
        title = extract_title(row['message_text'])
        company = channel_to_company.get(row['channel_username'], 'Unknown')
        sheet[f'B{current_row}'] = title
        sheet[f'C{current_row}'] = company
        date_str = datetime.now().strftime('%m/%d/%Y')
        date_obj = datetime.strptime(date_str, '%m/%d/%Y')
        sheet[f'D{current_row}'] = date_obj.month  # Заполнение месяца
        sheet[f'E{current_row}'] = date_obj.year  # Заполнение года
        sheet[f'F{current_row}'] = date_obj
        sheet[f'F{current_row}'].style = style_name
        sheet[f'AB{current_row}'] = row['message_text']
        sheet[f'AC{current_row}'] = f"https://t.me/{row['channel_username']}/{row['message_id']}"  # Ссылка на сообщение
        sheet[f'AD{current_row}'] = f"https://t.me/{row['channel_username']}"  # Ссылка на канал

        # Заполнение тегов
        for col_idx, prob in enumerate(probabilities, start=7):  # Столбцы G (7) - AA (27)
            if prob > 0.49:
                sheet.cell(row=current_row, column=col_idx).value = '+'

        current_row += 1

        # Обновление базы данных
        update_query = f"UPDATE messages SET analized = 1 WHERE id = {row['id']}"
        conn.execute(update_query)

# Сохранение изменений в Excel
wb.save(xlsx_path)

# Коммит изменений в базе данных
conn.commit()
conn.close()
